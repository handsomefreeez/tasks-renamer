# Powered By freeez
# 学委助手v0.1 根据文件名中的部分信息快速批量格式化文件名，并生成提交情况的表格

# 待完善：
# 完整的异常捕获
# 界面美化
# 存储历史数据
# 一次操作多个文件夹

import tkinter as tk
from tkinter import filedialog
from tkinter.constants import LEFT
import openpyxl
import os


global PATH  # 存储整理目录路径
global STUXLSXPATH  # 存储学生信息表格路径
global RENAME  # 用户是否使用重命名模式标记
global NEWXLSX  # 用户是否使用新建表格标记


PATH = ''
STUXLSXPATH = ''
RENAME = False
NEWXLSX = False


def Start():
    '''开始整理按钮的主要处理部分'''

    # 读取含有学号和姓名的excel表格，并把学号和姓名对应存储到两个数组中
    # （col1list、col2list），声明一个与学生和学号数组等长的status
    # 数组，来存储提交状态，缺陷是未做读取失败的处理
    wb = openpyxl.load_workbook(STUXLSXPATH)
    sheetnames = wb.sheetnames
    ws = wb[sheetnames[0]]
    status = [0] * ws.max_row
    col1list = []
    col2list = []
    col1_cell_list = list(ws.columns)[0]
    col2_cell_list = list(ws.columns)[1]
    for cell in col1_cell_list:
        col1list.append(cell.value)
    for cell in col2_cell_list:
        col2list.append(cell.value)

    # 在用户选择重命名文件时，进行文件重命名，无论用户是否选择重命名，
    # 该块都会统计每个学生是否已经提交，并存储在status数组的对应元素中
    for root, dirs, files in os.walk(PATH):
        for i in range(0, len(files)):
            dot = files[i].rfind('.')  # 切片出文件拓展名
            ends = files[i][dot:]
            for j in range(0, len(col1list)):
                if (files[i].find(col1list[j]) != -1):
                    if (RENAME.get() == 1):
                        os.rename(
                            root+'/'+files[i], root+'/'+str(col2list[j]) + col1list[j] + ends)  # 重命名为目录+/+文件名
                    status[j] = 1
                if (files[i].find(str(col2list[j])) != -1):
                    if (RENAME.get() == 1):
                        os.rename(
                            root+'/'+files[i], root+'/'+str(col2list[j]) + col1list[j] + ends)
                    status[j] = 1

    # 如果用户选择了生成提交状态表格，该代码块会将之前的学生学号表格复
    # 制到当前目录下，并在下一列写入是否提交
    if (NEWXLSX.get() == 1):
        for i in range(0, len(col1list)):
            if (status[i] == 0):
                var = '未交'
            else:
                var = '已交'
            ws.cell(row=i + 1, column=3, value=var)
            wb.save(root+'/'+'作业完成情况.xlsx')

    # 居中显示完成窗口（代码重用较差，暂未修改）
    helpwindow = tk.Tk()
    helpwindow.title('')
    curWidth = 250
    curHight = 30
    scn_w, scn_h = helpwindow.maxsize()
    cen_x = (scn_w - curWidth) / 2
    cen_y = (scn_h - curHight) / 2
    size_xy = '%dx%d+%d+%d' % (curWidth, curHight, cen_x, cen_y)
    helpwindow.geometry(size_xy)
    l = tk.Label(helpwindow, text='完成！', justify=LEFT)
    l.pack()


def Check():
    '''做正式开始整理前的合法性检查，如果合法，则开始整理，否则弹出窗口并终止'''

    # 检查路径是否合法
    if (len(PATH) < 1 and len(STUXLSXPATH) < 1):
        # 不合法弹出窗口提示用户
        helpwindow = tk.Tk()
        helpwindow.title('')
        curWidth = 250
        curHight = 30
        scn_w, scn_h = helpwindow.maxsize()
        cen_x = (scn_w - curWidth) / 2
        cen_y = (scn_h - curHight) / 2
        size_xy = '%dx%d+%d+%d' % (curWidth, curHight, cen_x, cen_y)
        helpwindow.geometry(size_xy)
        l = tk.Label(helpwindow, text='路径选择不完整，请重新设置两个路径', justify=LEFT)
        l.pack()

    # 两个模式都未选择时，整理操作无效
    elif (RENAME.get() == False and NEWXLSX.get() == False):
        # 不合法弹出窗口
        helpwindow = tk.Tk()
        helpwindow.title('')
        curWidth = 250
        curHight = 30
        scn_w, scn_h = helpwindow.maxsize()
        cen_x = (scn_w - curWidth) / 2
        cen_y = (scn_h - curHight) / 2
        size_xy = '%dx%d+%d+%d' % (curWidth, curHight, cen_x, cen_y)
        helpwindow.geometry(size_xy)
        l = tk.Label(helpwindow, text='未选择整理方式', justify=LEFT)
        l.pack()

    # 一切正常，开始执行整理
    else:
        Start()


def GetStuXlsx():
    '''获取学生信息表格的路径'''
    global STUXLSXPATH
    StuXlsxTuple = tk.filedialog.askopenfilenames(
        filetypes=[('Excle 文件（仅支持xlsx）', '.xlsx')])
    STUXLSXPATH = StuXlsxTuple[0]


def GetPath():
    '''获取要整理的目录的路径'''
    global PATH
    PATH = tk.filedialog.askdirectory()


def Help():
    '''弹出说明窗口'''
    helpwindow = tk.Tk()
    helpwindow.title('说明')
    curWidth = 600
    curHight = 500
    scn_w, scn_h = helpwindow.maxsize()
    cen_x = (scn_w - curWidth) / 2
    cen_y = (scn_h - curHight) / 2
    size_xy = '%dx%d+%d+%d' % (curWidth, curHight, cen_x, cen_y)
    helpwindow.geometry(size_xy)
    l = tk.Label(helpwindow, text='''
                选择名单文件：名单文件是一个excel表格，第一列填写\n
                姓名，第二列填写学号，切记不可包含表头，并且学号列\n
                和姓名列等长，否则会报错\n\n
                选择整理目录：选择一个含有要整理的作业的文件夹。\n\n
                重命名找到的作业：该选项被勾选时，在文件夹里面发现\n
                名单中包含的学号或者姓名时，会将其重命名成“学号+姓\n
                名”的格式。\n\n
                生成提交情况表格：该选项被勾选时，会在选择的文件夹\n
                中创建一个excel表格，其中第一列为姓名，第二列为学\n
                号，第三列为已交或未交。\n\n
                祝使用愉快。（目前不支持WPS，且目录中不能包含文件夹）
                ''', justify=LEFT)
    l.pack(side=LEFT)


def main():
    '''主函数'''
    # 主窗口居中显示
    window = tk.Tk()
    window.title('')
    curWidth = 150
    curHight = 300
    scn_w, scn_h = window.maxsize()
    cen_x = (scn_w - curWidth) / 2
    cen_y = (scn_h - curHight) / 2
    size_xy = '%dx%d+%d+%d' % (curWidth, curHight, cen_x, cen_y)
    window.geometry(size_xy)

    # 占位lable
    l = tk.Label(window, text=' ')
    l.pack()

    # 说明按钮
    helpbutton = tk.Button(window, text='说明', width=15, height=1, command=Help)
    helpbutton.pack()

    # 获取学生名单表格路径按钮
    getstubutton = tk.Button(window, text='选择名单文件',
                             width=15, height=1, command=GetStuXlsx)
    getstubutton.pack()

    # 获取要整理的目录路径
    getpathbutton = tk.Button(window, text='选择整理目录',
                              width=15, height=1, command=GetPath)
    getpathbutton.pack()

    # 主函数按钮，开始检查并执行
    mainbutton = tk.Button(window, text='开始整理', width=15,
                           height=1, command=Check)
    mainbutton.pack()

    # 复选框，检测用户是否启用了重命名功能
    global RENAME
    RENAME = tk.IntVar()
    renamecheck = tk.Checkbutton(
        window, text='重命名找到的作业', variable=RENAME, onvalue=1, offvalue=0)
    renamecheck.pack()

    # 复选框，检测用户是否启用了生成提交情况表格功能
    global NEWXLSX
    NEWXLSX = tk.IntVar()
    newxlsxcheck = tk.Checkbutton(
        window, text='生成提交情况表格', variable=NEWXLSX, onvalue=1, offvalue=0)
    newxlsxcheck.pack()

    # 签名信息lable
    l = tk.Label(window, text='V0.1 Powered By freeez')
    l.pack()

    window.mainloop()


if(__name__ == '__main__'):
    main()
